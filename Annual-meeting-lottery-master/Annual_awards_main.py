# coding:utf-8
import string
import numpy
import sys
import xlrd
import pygame
import random

from openpyxl.styles.colors import WHITE
from pygame.locals import *
from openpyxl import Workbook
import pygame

from j import test

fullScreen = True
mode = 0
mouse_scan = [(1031, 498, 116, 89), (1031, 436, 91, 68)]
pos_scan = [(683, 141, 683, 630), (650, 130, 650, 540), ()]
winprize = ['特等奖', '一等奖', '二等奖', '福利']


def get_name_list_from_excel(file_name):
    '''"解析 人员.xlsx 文件，得到人员名单列表"'''
    name_list = []
    excelFile = xlrd.open_workbook(file_name)
    sheet = excelFile.sheet_by_name('Sheet1')
    print(sheet.name, sheet.nrows, sheet.ncols)
    job_num = sheet.cell(0, 0).value
    job_name = sheet.cell(0, 1).value
    job_sex = sheet.cell(0, 2).value
    job_contact = sheet.cell(0, 3).value
    for row in range(1, sheet.nrows):
        job_num = sheet.cell(row, 0).value
        job_name = sheet.cell(row, 1).value
        job_sex = sheet.cell(row, 2).value
        job_contact = sheet.cell(row, 3).value
        # print job_num, job_name
        name_list.append((job_num, job_name, job_sex, job_contact))
    return job_num, job_name, job_sex, job_contact, name_list


def handle_mouse_event(index, pause_flag):
    if pygame.mouse.get_pressed()[0]:
        x, y = pygame.mouse.get_pos()
        x -= m.get_width() / 2
        y -= m.get_height() / 2
        if mouse_scan[mode][0] + mouse_scan[mode][2] > x > mouse_scan[mode][0] and mouse_scan[mode][1] + \
                mouse_scan[mode][3] > y > mouse_scan[mode][1]:
            pause_flag = True if pause_flag == False else False
            if not pause_flag:
                pygame.mixer.music.play()
                del name_list[index]
                # print len(name_list)
                # show_name_list(name_list)
            else:
                pygame.mixer.music.stop()

    return index, pause_flag


def show_name_list(name_list):
    """
    调试接口，输出当前名单
    """
    for index in range(0, len(name_list)):
        str = "%s %s %s %s" % (name_list[index][0], name_list[index][1], name_list[index][2], name_list[index][3])
        print(str)


if __name__ == "__main__":
    # test()
    # if res:
    test()
    name_list = []
    work = []
    print("请选择人员添加方式  1:手动输入,2:excel文件导入")
    a = int(input("选择:"))
    if a == 1:
        # 创建一个新的Excel工作簿
        workbook = Workbook()

        # 获取默认的工作表
        worksheet = workbook.active

        # 获取抽奖人数
        num_of_entries = int(input("请输入抽奖人数："))

        # 循环输入抽奖人员信息
        for i in range(num_of_entries):
            job_num = input("请输入第{}个抽奖人员的学号：".format(i + 1))
            job_name = input("请输入第{}个抽奖人员的姓名：".format(i + 1))
            job_sex = input("请输入第{}个抽奖人员的性别：".format(i + 1))
            job_contact = input("请输入第{}个抽奖人员的联系方式：".format(i + 1))
            # 将抽奖人员信息写入Excel表格中的每一行
            worksheet.append([job_num, job_name, job_sex, job_contact])
            name_list.append([job_num, job_name, job_sex, job_contact])
            work.append([job_num, job_name, job_sex, job_contact])
        # 保存Excel文件
        filename = input("请输入要保存的Excel文件名（包含文件扩展名）：")
        workbook.save(filename)

        print("数据已成功导入Excel文件！")
        # job_num, job_name, job_sex, job_contact, name_list = get_name_list_from_excel(r'filename')
        print(len(name_list))
        show_name_list(name_list)
    if a == 2:
        job_num, job_name, job_sex, job_contact, name_list = get_name_list_from_excel(r'name_file.xlsx')
        # job1, job2, job3, job4, work = get_name_list_from_excel(r'name_file.xlsx')
        filename1 = input("请输入文件名（包含文件扩展名）：")
        job1, job2, job3, job4, work = get_name_list_from_excel(filename1)
        print(len(name_list))
        show_name_list(name_list)

    pygame.init()
    bg = 'bg.png'
    mg = 'gc_cz.png'

    if fullScreen:
        mode = 0
        bg = 'bg_1366x768.png'
        screen = pygame.display.set_mode((1366, 768), FULLSCREEN, 32)
    else:
        mode = 1
        bg = 'bg.png'
        screen = pygame.display.set_mode((1340, 670), 0, 32)

    pygame.display.set_caption("Annual meeting lottery")

    b = pygame.image.load(bg).convert()
    m = pygame.image.load(mg).convert_alpha()
    screen.blit(b, (0, 0))
    screen.blit(m, (0, 0))

    pygame.mixer.init()
    pygame.mixer.music.load('chicken.mp3')
    # pygame.mixer.music.set_volume(0.5)
    pygame.mixer.music.play()
    # pygame.init()
    while True:
        print("是否选择退出 0:退出 1:继续")
        ans = int(input())
        if ans == 0:
            sys.exit()
        print('请选择奖品项目: 0:特等奖, 1:一等奖, 2:二等奖, 3:福利')
        h = int(input('请输入你的选择:'))
        print("是否选择重置抽奖名单1:重置,2:不重置")
        p = int(input('输入你的选择:'))
        if p == 1:
            name_list = work
        print('请输入奖项数目:')
        pirze_num = int(input())
        index = random.randint(0, len(name_list) - 1)
        pause_flag = False
        i = 0
        pygame.init()
        bg = 'bg.png'
        mg = 'gc_cz.png'

        if fullScreen:
            mode = 0
            bg = 'bg_1366x768.png'
            screen = pygame.display.set_mode((1366, 768), FULLSCREEN, 32)
        else:
            mode = 1
            bg = 'bg.png'
            screen = pygame.display.set_mode((1340, 670), 0, 32)

        pygame.display.set_caption("Annual meeting lottery")

        b = pygame.image.load(bg).convert()
        m = pygame.image.load(mg).convert_alpha()
        screen.blit(b, (0, 0))
        screen.blit(m, (0, 0))

        pygame.mixer.init()
        pygame.mixer.music.load('chicken.mp3')
        # pygame.mixer.music.set_volume(0.5)
        pygame.mixer.music.play()
        winner = []
        while i < 2 * pirze_num:
            for event in pygame.event.get():
                if event.type == pygame.KEYDOWN and event.key == pygame.K_ESCAPE:
                    pygame.quit()
                    sys.exit(0)
                elif event.type == MOUSEBUTTONDOWN:
                    i = i + 1
                    index, pause_flag = handle_mouse_event(index, pause_flag)
                    if i % 2 == 1:
                        winner.append(
                            (name_list[index][0], name_list[index][1], name_list[index][2], name_list[index][3]))
            screen.blit(b, (0, 0))
            x, y = pygame.mouse.get_pos()
            x -= m.get_width() / 2
            y -= m.get_height() / 2
            pygame.mouse.set_visible(False)
            screen.blit(m, (x, y))

            if not pause_flag:
                index = random.randint(0, len(name_list) - 1)
                if not pygame.mixer.music.get_busy():
                    pygame.mixer.music.play()

            text_context = '%s ' \
                           '%s  ' \
                           '%s ' \
                           '%s' % (name_list[index][0], name_list[index][1], name_list[index][2], name_list[index][3])
            # print text_context

            font = pygame.font.Font("simhei.ttf", 80)
            text_obj = font.render(text_context, True, (255, 255, 255), (0, 0, 0))
            text_pos = text_obj.get_rect()
            text_pos.center = (pos_scan[mode][0], pos_scan[mode][1])
            screen.blit(text_obj, text_pos)

            text_prize = '恭喜获得奖项'':''%s' % (winprize[h])
            font = pygame.font.Font("simhei.ttf", 30)
            text_obj = font.render(text_prize, True, (255, 255, 255), (255, 0, 0))
            text_pos = text_obj.get_rect()
            text_pos.center = (pos_scan[mode][2], pos_scan[mode][3])
            screen.blit(text_obj, text_pos)

            pygame.display.update()
        pygame.quit()
        print("该轮获奖者名单")
        show_name_list(winner)
        pygame.init()
        font = pygame.font.Font("simhei.ttf", 30)
        screen2 = pygame.display.set_mode((1200, 600))
        bg2 = pygame.image.load("backup.jpg").convert_alpha()
        screen2.blit(bg2, (0, 0))
        game_title = font.render('Prizer_Winner', True, (255, 255, 255), (255, 0, 0))
        screen2.blit(game_title, (1200 // 2 - game_title.get_width() // 2, 150))

        for index in range(0, len(winner)):
            res = "%s %s %s %s" % (winner[index])
            text_obj = font.render(res, True, (255, 255, 255), (255, 0, 0))
            text_pos = text_obj.get_rect()
            text_pos.center = (200, index*100+30)
            screen2.blit(text_obj, text_pos)
            pygame.display.update()

        if pygame.mouse.get_pressed()[0]:
            pygame.quit()
        # for event in pygame.event.get():
        #     if event.type == pygame.QUIT:
        #         pygame.quit()
        #         raise SystemExit
        #     if event.type == pygame.MOUSEBUTTONDOWN:
        #         pygame.quit()

        # for event in pygame.event.get():
        #     if event.type == pygame.KEYDOWN and event.key == pygame.K_ESCAPE:
        #         pygame.quit()
        #         sys.exit(0)
        #     elif event.type == MOUSEBUTTONDOWN:
        #         pygame.quit()
